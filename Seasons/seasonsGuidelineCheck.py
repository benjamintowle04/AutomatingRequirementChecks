from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import time


wb_requirementsCheck = load_workbook(r"C:/DSSO/AutomatingRequirementChecks/Seasons/Shifts.xlsx")


wb_guidelines = load_workbook(
    r"C:/DSSO/AutomatingRequirementChecks/Seasons/seasons_Guidelines.xlsx"
)


class Employee:
    def __init__(
        self,
        last_name,
        first_name,
        shift_count,
        hours,
        supervisor,
        missing_reqs,
        schedule,
        schedule_valid,
        notes
    ):
        self.last_name = last_name
        self.first_name = first_name
        self.shift_count = shift_count
        self.hours = hours
        self.supervisor = supervisor
        self.missing_reqs = missing_reqs
        self.schedule = schedule
        self.schedule_valid = schedule_valid
        self.notes = notes

    def __str__(self):
        return (
            f"Employee({self.first_name} {self.last_name}, "
            f"Hours: {self.hours}, "
            f"Missing Requirements: {self.missing_reqs})"
        )


class Shift:
    def __init__(self, day, start_time, end_time, hours, station, group):
        self.day = day
        self.start_time = start_time
        self.end_time = end_time
        self.hours = hours
        self.station = station
        self.group = group


class Guideline:
    def __init__(self, requirement_list, exceptions, min_hours):
        self.requirement_list = requirement_list
        self.exceptions = exceptions
        self.min_hours = min_hours


class Requirement:
    def __init__(self, requirement_name, requirement_type, excused, min_shifts):
        self.requirement_name = requirement_name
        self.requirement_type = requirement_type
        self.excused = excused
        self.min_shifts = min_shifts


class TimeRequirement:
    def __init__(self, min_start, max_start, min_end, max_end):
        self.min_start = min_start
        self.max_start = max_start
        self.min_end = min_end
        self.max_end = max_end

class TimeInNeed:
    def __init__(self, time_period, min_start, max_start, min_end, max_end):
        self.time_period = time_period
        self.min_start = min_start
        self.max_start = max_start
        self.min_end = min_end
        self.max_end = max_end

class DayRequirement:
    def __init__(self, days_list):
        self.days_list = days_list


class ShiftTypeRequirement:
    def __init__(self, types_list):
        self.types_list = types_list


class Exceptions:
    def __init__(self, exception_note, excuse_req_list, requirement_excused, min_shifts):
        self.exception_note = exception_note
        self.excuse_req_list = excuse_req_list
        self.requirement_excused = requirement_excused
        self.min_shifts = min_shifts


def main():
    # iterators
    i = 0
    j = 0
    k = 2

    # Initialize Worksheets
    ws_schedule = wb_requirementsCheck["Sheet1"]
    
    #So we don't continuously create more sheets than needed
    if "Call Sheet" in wb_requirementsCheck.sheetnames:
        ws_callsheet = wb_requirementsCheck["Call Sheet"]
        
        #Clear the existing call sheet 
        ws_callsheet.delete_rows(1, ws_callsheet.max_row)
        ws_callsheet.delete_cols(1, ws_callsheet.max_column)
        
    else:
        wb_requirementsCheck.create_sheet("Call Sheet")
        ws_callsheet = wb_requirementsCheck["Call Sheet"]
        
    ws_requirements = wb_guidelines["Seasons Requirements"]
    ws_requirements_supervisor = wb_guidelines["Seasons Supervisor Requirements"]
    ws_times_in_need = wb_guidelines["Seasons Times in Need"]
    

    employeeList = []

    #Initialize Guidelines 
    seasons_regular_guideline = getFacilityGuidelineRegular(ws_requirements, ws_times_in_need)
    seasons_supervisor_guideline = getFacilityGuidelineSupervisor(ws_requirements_supervisor, ws_times_in_need)

    # Initialize Callsheet Column Headers
    ws_callsheet["A1"].value = "Last"
    ws_callsheet["B1"].value = "First"
    ws_callsheet["C1"].value = "# Shifts"
    ws_callsheet["D1"].value = "Hours"
    ws_callsheet["E1"].value = "SUP?"
    ws_callsheet["F1"].value = "Missing Requirement(s)"
    ws_callsheet["G1"].value = "Notes"

    # Initialize Employees' first names, last names, and supervisor status
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("A"),
        max_col=column_index_from_string("A"),
        min_row=2,
    ):
        full_name = row[0].value
        if full_name:
            supervisor = "No"
            if "(SUP)" in full_name:
                supervisor = "Yes"
                # Remove (SUP) from first name
                substring_to_remove = " (SUP)"
                full_name = full_name.replace(substring_to_remove, "")
            full_name = full_name.rstrip()

            employee = Employee(
                getLastName(full_name),
                getFirstName(full_name),
                0,
                0.0,
                supervisor,
                [],
                [],
                True,
                ""
            )
            employeeList.append(employee)

    # initialize Employees' shift count
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("F"),
        max_col=column_index_from_string("F"),
        min_row=3,
        max_row=ws_schedule.max_row - 1  # Exclude the last row
    ):
        val = row[0].value
        if val and isinstance(val, str):
            if "Count: " in val:
                employeeList[i].shift_count = getShiftCount(val)
                i += 1

    # initialize Employees' Hours
    employeeCount = 0
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("H"),
        max_col=column_index_from_string("H"),
        min_row=3,
        max_row=ws_schedule.max_row - 1  # Exclude the last row
    ):
        val = row[0].value
        if val:
            if j == employeeList[employeeCount].shift_count:
                employeeList[employeeCount].hours = float(val)
                j = 0
                employeeCount += 1
            else:
                j += 1

    # Initialize employees schedule
    for employee in employeeList:
        for colB, colC, colD, colE, colF, colG, colH in zip(
            ws_schedule.iter_cols(min_col=2, max_col=2, values_only=True),
            ws_schedule.iter_cols(min_col=3, max_col=3, values_only=True),
            ws_schedule.iter_cols(min_col=4, max_col=4, values_only=True),
            ws_schedule.iter_cols(min_col=5, max_col=5, values_only=True),
            ws_schedule.iter_cols(min_col=6, max_col=6, values_only=True),
            ws_schedule.iter_cols(min_col=7, max_col=7, values_only=True),
            ws_schedule.iter_cols(min_col=8, max_col=8, values_only=True),
        ):
            for B, C, D, E, F, G, H in zip(colB, colC, colD, colE, colF, colG, colH):
                if G:
                    if employee.last_name in G and employee.first_name in G:

                        employee.schedule.append(
                            Shift(
                                get_day_of_week(B),
                                convert_to_time(C),
                                convert_to_time(D),
                                float(H),
                                F,
                                E
                            )
                        )


    guidelineCheck(seasons_regular_guideline, seasons_supervisor_guideline, employeeList)
    filterList(employeeList)

    # Add Rows to new worksheet
    for employee in employeeList:
        ws_callsheet[f"A{k}"] = employee.last_name
        ws_callsheet[f"B{k}"] = employee.first_name
        ws_callsheet[f"C{k}"] = employee.shift_count
        ws_callsheet[f"D{k}"] = employee.hours
        ws_callsheet[f"E{k}"] = employee.supervisor
        ws_callsheet[f"F{k}"] = ", ".join(employee.missing_reqs)
        ws_callsheet[f"G{k}"] = employee.notes

        k += 1

    wb_requirementsCheck.save("Shifts.xlsx")
    print("FINISHED")

    
    
def countRows(ws):
    count = 0
    for row in ws.iter_rows():
        count += 1
    return count

# Helper function to convert the day of the year into its respective day of the week
def get_day_of_week(date_obj):
    # date_obj = datetime.datetime.strptime(date_str, "%m/%d/%Y")
    day_of_week_num = date_obj.weekday()
    days_of_week = [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ]
    day_of_week = days_of_week[day_of_week_num]
    return day_of_week


# Helper function to return integer value of the column (Used to iterate through a column)
def column_index_from_string(column):
    index = 0
    for char in column.upper():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


# Retrieve last name from the full name of employee
def getLastName(employeeName):
    try:
        employeeName = employeeName[10:]
        last_name, first_name = employeeName.split(", ")
        return last_name
    except ValueError:
        return "Invalid format. Please use 'Last, First'."


# Retrieve first name from the full name of employee
def getFirstName(employeeName):
    try:
        employeeName = employeeName[10:]
        last_name, first_name = employeeName.split(", ")
        return first_name
    except ValueError:
        return "Invalid format. Please use 'Last, First'."


# Retrieve the number of shifts the employee is working
def getShiftCount(cellValue):
    stringNumber = cellValue[7:]
    return int(stringNumber)


def getFacilityGuidelineRegular(ws_reqs, ws_times):
    req_list = []
    excused_req_list = []
    shift_excused = None
    exceptions = None

    for row in ws_reqs.iter_rows(min_row=2):
        # Initialize column values from sheet
        row_number = row[0].row
        req_name = row[column_index_from_string("A") - 1].value
        req_type = row[column_index_from_string("B") - 1].value
        days = decodeDaysCheckboxCell(row_number, ws_reqs)
        min_start = row[column_index_from_string("D") - 1].value
        max_start = row[column_index_from_string("E") - 1].value
        min_end = row[column_index_from_string("F") - 1].value
        max_end = row[column_index_from_string("G") - 1].value
        cell_value = row[column_index_from_string("J") - 1].value
        min_exception_shifts = row[column_index_from_string("L") - 1].value
        exception_note = row[column_index_from_string("M") - 1].value
        shift_types_list = decodeShiftTypeCheckboxCell(row_number, ws_reqs)

        if not req_name:
            continue
        
        if cell_value.lower() == "y":
            excused = True
        else:
            excused = False

        if req_type == "Day":
            day_req_type = DayRequirement(days)
            day_req = Requirement(req_name, day_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = day_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
                
            req_list.append(day_req)

        elif req_type == "Time":
            time_req_type = TimeRequirement(min_start, max_start, min_end, max_end)
            time_req = Requirement(req_name, time_req_type, excused, 1)

            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = time_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)

                
            req_list.append(time_req)

        elif req_type == "Shift Type":
            shiftType_req_type = ShiftTypeRequirement(shift_types_list)
            shiftType_req = Requirement(req_name, shiftType_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = shiftType_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
                
            req_list.append(shiftType_req)

        else:
            print("Error: Invalid Input")

    return Guideline(req_list, exceptions, 8)


def getFacilityGuidelineSupervisor(ws_reqs, ws_times):
    req_list = []
    excused_req_list = []
    shift_excused = None
    exceptions = None

    for row in ws_reqs.iter_rows(min_row=2):
        # Initialize column values from sheet
        row_number = row[0].row
        req_name = row[column_index_from_string("A") - 1].value
        req_type = row[column_index_from_string("B") - 1].value
        days = decodeDaysCheckboxCell(row_number, ws_reqs)
        min_start = row[column_index_from_string("D") - 1].value
        max_start = row[column_index_from_string("E") - 1].value
        min_end = row[column_index_from_string("F") - 1].value
        max_end = row[column_index_from_string("G") - 1].value
        cell_value = row[column_index_from_string("J") - 1].value
        min_exception_shifts = row[column_index_from_string("L") - 1].value
        exception_note = row[column_index_from_string("M") - 1].value
        shift_types_list = decodeShiftTypeCheckboxCell(row_number, ws_reqs)

        if not req_name:
            continue
        
        if cell_value.lower() == "y":
            excused = True
        else:
            excused = False


        if req_type == "All SUP":            
            shiftType_req_type = ShiftTypeRequirement(["All SUP"])
            shiftType_req = Requirement(req_name, shiftType_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = shiftType_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)    
            req_list.append(shiftType_req)
            
        elif req_type == "One SUP":
            shiftType_req_type = ShiftTypeRequirement(["One SUP"])
            shiftType_req = Requirement(req_name, shiftType_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = shiftType_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
            req_list.append(shiftType_req)
            
        elif req_type == "Multiple SUPs":
            shiftType_req_type = ShiftTypeRequirement(["Multiple SUPs"])
            shiftType_req = Requirement(req_name, shiftType_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = shiftType_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
                
            req_list.append(shiftType_req)
            
        elif req_type == "Day":
            day_req_type = DayRequirement(days)
            day_req = Requirement(req_name, day_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = day_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
                
            req_list.append(day_req)

        elif req_type == "Time":
            time_req_type = TimeRequirement(min_start, max_start, min_end, max_end)
            time_req = Requirement(req_name, time_req_type, excused, 1)

            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = time_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)

                
            req_list.append(time_req)

        elif req_type == "Shift Type":
            shiftType_req_type = ShiftTypeRequirement(shift_types_list)
            shiftType_req = Requirement(req_name, shiftType_req_type, excused, 1)
            if excused and min_exception_shifts >= 1:
                excused_req_list = decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times)
                shift_excused = shiftType_req
                exceptions = Exceptions(exception_note, excused_req_list, shift_excused, min_exception_shifts)
                
            req_list.append(shiftType_req)

        else:
            print("Error: Invalid Input")

    return Guideline(req_list, exceptions, 13)


def decodeExceptionsCheckBoxCell(row_number, ws_reqs, ws_times):
    shift_list = []
    times_in_need = getTimesInNeed(ws_times)
    row_number *= 10
    breakfast = ws_reqs.cell(row=row_number, column=11).value
    lunch = ws_reqs.cell(row=row_number + 1, column=11).value
    dinner = ws_reqs.cell(row=row_number + 2, column=11).value
    late = ws_reqs.cell(row=row_number + 3, column=11).value
    na = ws_reqs.cell(row=row_number + 4, column=11).value
    

    if na:
        return []
     

    if breakfast:
        for item in times_in_need:
            if item.time_period == "Breakfast":
                req_type = TimeRequirement(item.min_start, item.max_start, item.min_end, item.max_end)
                req = Requirement("Breakfast", req_type, False, 1)
                shift_list.append(req)

    if lunch:
        for item in times_in_need:
            if item.time_period == "Lunch":
                req_type = TimeRequirement(item.min_start, item.max_start, item.min_end, item.max_end)
                req = Requirement("Lunch", req_type, False, 1)
                shift_list.append(req)

    if dinner:
        for item in times_in_need:
            if item.time_period == "Dinner":
                req_type = TimeRequirement(item.min_start, item.max_start, item.min_end, item.max_end)
                req = Requirement("Dinner", req_type, False, 1)
                shift_list.append(req)

    if late:
        for item in times_in_need:
            if item.time_period == "Late":
                req_type = TimeRequirement(item.min_start, item.max_start, item.min_end, item.max_end)
                req = Requirement("Late", req_type, False, 1)
                shift_list.append(req)

    return shift_list


def decodeDaysCheckboxCell(row_number, ws):
    day_list = []
    row_number *= 10
    monday = ws.cell(row=row_number, column=3).value
    tuesday = ws.cell(row=row_number + 1, column=3).value
    wednesday = ws.cell(row=row_number + 2, column=3).value
    thursday = ws.cell(row=row_number + 3, column=3).value
    friday = ws.cell(row=row_number + 4, column=3).value
    saturday = ws.cell(row=row_number + 5, column=3).value
    sunday = ws.cell(row=row_number + 6, column=3).value
    na = ws.cell(row=row_number + 7, column=3).value

    if na:
        return []

    if monday:
        day_list.append("Monday")
    if tuesday:
        day_list.append("Tuesday")
    if wednesday:
        day_list.append("Wednesday")
    if thursday:
        day_list.append("Thursday")
    if friday:
        day_list.append("Friday")
    if saturday:
        day_list.append("Saturday")
    if sunday:
        day_list.append("Sunday")

    return day_list


def decodeShiftTypeCheckboxCell(row_number, ws):
    shift_list = []
    row_number *= 10
    dish = ws.cell(row=row_number, column=8).value
    dining = ws.cell(row=row_number + 1, column=8).value
    greeter = ws.cell(row=row_number + 2, column=8).value
    beverages = ws.cell(row=row_number + 3, column=8).value
    boh = ws.cell(row=row_number + 4, column=8).value
    na = ws.cell(row=row_number + 5, column=8).value

    if na:
        return []

    if dish:
        shift_list.append("Dish Room")
    if dining:
        shift_list.append("Dining Room")
    if greeter:
        shift_list.append("Greeter")
    if beverages:
        shift_list.append("Beverages")
    if boh:
        shift_list.append("Back of House")

    return shift_list


def guidelineCheck(guideline, sup_guideline, employeeList):
    for employee in employeeList:
        if employee.supervisor == "Yes":
            req_list = sup_guideline.requirement_list
            if sup_guideline.exceptions:
                exceptions = sup_guideline.exceptions
                req_excused = exceptions.requirement_excused
            else:
                exceptions = None
            min_hours = sup_guideline.min_hours

        else:
            req_list = guideline.requirement_list
            exceptions = guideline.exceptions
            if exceptions:
                req_excused = exceptions.requirement_excused
            min_hours = guideline.min_hours

        employee.schedule_valid = True #Employees are innocent until proven guilty
        for req in req_list:
            req_type = req.requirement_type
            req_name = req.requirement_name
            if not (
                meetsRequirement(
                    employee, req_name, req_type, req.min_shifts, min_hours
                )
            ):
                # Check for excuses in the requirement
                if exceptions and req == req_excused:
                    if excuseRequirement(employee, exceptions, req):
                        employee.notes = exceptions.exception_note
                        
                        if req.requirement_name in employee.missing_reqs:
                            employee.missing_reqs.remove(req.requirement_name)

                            
                    else :
                        employee.schedule_valid = False
                
                else:
                    employee.schedule_valid = False
    return employeeList


def excuseRequirement(employee, exceptions, req):
    req_list = (
        exceptions.excuse_req_list
    )  # If any of these requirements are met, it counts towards 1 shift in the min_shifts
    meets_req_count = 0
    min_shifts = exceptions.min_shifts

    for req in req_list:
        meets_req_count += countMetRequirements(employee, req.requirement_type)

    if meets_req_count >= min_shifts:
        return True

    else:
        return False


def countMetRequirements(employee, requirement_type):
    shift_count = 0

    if isinstance(requirement_type, DayRequirement):
        days = requirement_type.days_list
        for shift in employee.schedule:
            for day in days:
                if shift.day == day:
                    shift_count += 1

    elif isinstance(requirement_type, ShiftTypeRequirement):
        types = requirement_type.types_list
        for shift in employee.schedule:
            for shiftType in types:
                if shift.station == shiftType:
                    shift_count += 1

    elif isinstance(requirement_type, TimeRequirement):
        min_start = requirement_type.min_start
        min_end = requirement_type.min_end
        max_start = requirement_type.max_start
        max_end = requirement_type.max_end

        # Set necessary upper/lower bounds of times for values that are null
        if min_start == None:
            min_start = convert_to_time("12:00:00 AM")

        if max_start == None:
            max_start = convert_to_time("11:59:00 PM")

        if min_end == None:
            min_end = convert_to_time("12:00:00 AM")

        if max_end == None:
            max_end = convert_to_time("11:59:00 PM")

        # Sift through schedule and make sure at least 1 shift is in range
        for shift in employee.schedule:
            if (
                shift.start_time >= min_start
                and shift.start_time <= max_start
                and shift.end_time >= min_end
                and shift.end_time <= max_end
            ):
                shift_count += 1

    return shift_count


def meetsRequirement(
    employee, requirement_name, requirement_type, min_shifts, min_hours
):
    shift_count = 0
    meetsRequirement = False

    if isinstance(requirement_type, DayRequirement):
        days = requirement_type.days_list
        for shift in employee.schedule:
            for day in days:
                if shift.day == day:
                    shift_count += 1
                    if shift_count == min_shifts:
                        meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append(requirement_name)

    elif isinstance(requirement_type, ShiftTypeRequirement):
        types = requirement_type.types_list
        if "All SUP" in types:
            types = ["Supervisor"]
            min_shifts = employee.shift_count
            
        elif "One SUP" in types:
            types = ["Supervisor"]
            min_shifts = 1
            
        elif "Multiple SUPs" in types:
            types = ["Supervisor"]
            min_shifts = 2

        for shift in employee.schedule:
            for shiftType in types:
                if employee.supervisor == "Yes" and shiftType != "Supervisor":
                    m_shifts = 1
                    if shift.group == shiftType:
                        shift_count += 1
                        if shift_count == m_shifts:
                            meetsRequirement = True
                else:
                    if shift.station == shiftType:
                        shift_count += 1
                        if shift_count == min_shifts:
                            meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append(requirement_name)

    elif isinstance(requirement_type, TimeRequirement):
        min_start = requirement_type.min_start
        min_end = requirement_type.min_end
        max_start = requirement_type.max_start
        max_end = requirement_type.max_end

        # Set necessary upper/lower bounds of times for values that are null
        if min_start == None or isinstance(min_start, str):
            min_start = convert_to_time("12:00:00 AM")

        if max_start == None or isinstance(max_start, str):
            max_start = convert_to_time("11:59:00 PM")

        if min_end == None or isinstance(min_end, str):
            min_end = convert_to_time("12:00:00 AM")

        if max_end == None or isinstance(max_end, str):
            max_end = convert_to_time("11:59:00 PM")
            

        # Sift through schedule and make sure at least 1 shift is in range
        for shift in employee.schedule:
            
            #If a shift ends later than 12 AM, we want its end time to be at max
            if shift.end_time >= convert_to_time("12:00:00 AM") and shift.end_time <= convert_to_time("3:00:00 AM"):
                shift.end_time = convert_to_time("11:59:00 PM")
                
                
            if (
                shift.start_time >= min_start
                and shift.start_time <= max_start
                and shift.end_time >= min_end
                and shift.end_time <= max_end
            ):
                shift_count += 1
                if shift_count == min_shifts:
                    meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append(requirement_name)

    if employee.hours < min_hours:
        meetsRequirement = False
        if not ("Under Hours" in employee.missing_reqs):
            employee.missing_reqs.append("Under Hours")

    return meetsRequirement


def filterList(employeeList):
    for employee in employeeList[:]:  # Iterate over a shallow copy of the list
        if employee.schedule_valid:
            employeeList.remove(employee)


def getTimesInNeed(ws):
    times_list = []
    
    if ws:
        for colA, colB, colC, colD, colE in zip(  
            ws.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True),
            ws.iter_cols(min_col=2, max_col=2, min_row=2, values_only=True),
            ws.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True),
            ws.iter_cols(min_col=4, max_col=4, min_row=2, values_only=True),
            ws.iter_cols(min_col=5, max_col=5, min_row=2, values_only=True)
        ):
            for A, B, C, D, E in zip(colA, colB, colC, colD, colE):
                if A: 
                    time_period = A
                    min_start = B
                    max_start = C
                    min_end = D
                    max_end = E
                    times_list.append(TimeInNeed(time_period, min_start, max_start, min_end, max_end))
                    
    return times_list
                    

def convert_to_time(time_input):
    # Check if the input is a datetime object
    if isinstance(time_input, datetime):
        datetime_obj = time_input
    else:
        # Define the datetime formats
        full_datetime_format = "%m/%d/%Y %I:%M:%S %p"
        time_only_format = "%I:%M:%S %p"

        try:
            # Try to parse the full date-time string
            datetime_obj = datetime.strptime(time_input, full_datetime_format)
        except ValueError:
            # If that fails, try to parse the time-only string
            datetime_obj = datetime.strptime(time_input, time_only_format)

    # Extract the time component from the datetime object
    time_struct = datetime_obj.time()

    return time_struct


if __name__ == "__main__":
    main()
