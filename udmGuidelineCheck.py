from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import time


wb_requirementsCheck = load_workbook(
    r"C:\DSSO\AutomatingRequirementChecks\SampleRequirementsCheckForBen.xlsx"
)


wb_guidelines = load_workbook(
    r"C:\DSSO\AutomatingRequirementChecks\UDM_Guidelines.xlsx"
)


class Employee:
    def __init__(
        self,
        last_name,
        first_name,
        shift_count,
        hours,
        supervisor,
        missing_reqs,
        schedule,
        schedule_valid,
    ):
        self.last_name = last_name
        self.first_name = first_name
        self.shift_count = shift_count
        self.hours = hours
        self.supervisor = supervisor
        self.missing_reqs = missing_reqs
        self.schedule = schedule
        self.schedule_valid = schedule_valid

    def __str__(self):
        return (
            f"Employee({self.first_name} {self.last_name}, "
            f"Hours: {self.hours}, "
            f"Missing Requirements: {self.missing_reqs})"
        )


class Shift:
    def __init__(self, day, start_time, end_time, hours, station):
        self.day = day
        self.start_time = start_time
        self.end_time = end_time
        self.hours = hours
        self.station = station


class Guideline:
    def __init__(self, requirement_list, exceptions, min_hours):
        self.requirement_list = requirement_list
        self.exceptions = exceptions
        self.min_hours = min_hours


class Requirement:
    def __init__(self, requirement_type, excused, min_shifts):
        self.requirement_type = requirement_type
        self.excused = excused
        self.min_shifts = min_shifts


class TimeRequirement:
    def __init__(self, min_start, max_start, min_end, max_end):
        self.min_start = min_start
        self.max_start = max_start
        self.min_end = min_end
        self.max_end = max_end


class DayRequirement:
    def __init__(self, days_list):
        self.days_list = days_list


class ShiftTypeRequirement:
    def __init__(self, types_list):
        self.types_list = types_list


class Exceptions:
    def __init__(self, excuse_req_list, requirement_excused, min_shifts):
        self.excuse_req_list = excuse_req_list
        self.requirement_excused = requirement_excused
        self.min_shifts = min_shifts


def main():
    # iterators
    i = 0
    j = 0
    k = 2

    # Initialize Worksheets
    ws_schedule = wb_requirementsCheck.active
    wb_requirementsCheck.create_sheet("Call Sheet")
    ws_callsheet = wb_requirementsCheck["Call Sheet"]
    ws_requirements = wb_guidelines.active

    employeeList = []

    # Requirement Types Defined
    weekend = DayRequirement(["Saturday", "Sunday"])
    dishroom = ShiftTypeRequirement(["Dish Room"])
    late = TimeRequirement(convert_to_time("7:00:00 PM"), None, None, None)

    # Requirements
    weekend_req = Requirement(weekend, False, 1)
    dish_req = Requirement(dishroom, False, 1)
    late_req = Requirement(late, False, 1)

    # Excuses
    breakfast = TimeRequirement(None, convert_to_time("8:00:00 AM"), None, None)
    lunch = TimeRequirement(
        convert_to_time("10:00:00 AM"), convert_to_time("12:00:00 PM"), None, None
    )
    breakfast_req = Requirement(breakfast, False, 1)
    lunch_req = Requirement(lunch, False, 1)
    exceptions = Exceptions([breakfast_req, lunch_req], weekend_req, 3)

    # High Level Guideline for UDM Regular Workers
    udm_regular_guideline = Guideline([weekend_req, dish_req, late_req], exceptions, 8)

    # Set up supervisor guidelines
    all_supervisor = ShiftTypeRequirement(["Supervisor"])
    all_supervisor_req = Requirement(all_supervisor, False, 1)
    udm_supervisor_guideline = Guideline([all_supervisor_req, weekend_req], None, 13)

    # Initialize Callsheet Column Headers
    ws_callsheet["A1"].value = "Last"
    ws_callsheet["B1"].value = "First"
    ws_callsheet["C1"].value = "# Shifts"
    ws_callsheet["D1"].value = "Hours"
    ws_callsheet["E1"].value = "SUP?"
    ws_callsheet["F1"].value = "Missing Requirement(s)"

    # Initialize Employees' first names, last names, and supervisor status
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("A"),
        max_col=column_index_from_string("A"),
        min_row=2,
    ):
        full_name = row[0].value
        if full_name:
            supervisor = "No"
            if "(SUP)" in full_name:
                supervisor = "Yes"
                # Remove (SUP) from first name
                substring_to_remove = " (SUP)"
                full_name = full_name.replace(substring_to_remove, "")

            employee = Employee(
                getLastName(full_name),
                getFirstName(full_name),
                0,
                0.0,
                supervisor,
                [],
                [],
                True,
            )
            employeeList.append(employee)

    # initialize Employees' shift count
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("F"),
        max_col=column_index_from_string("F"),
        min_row=3,
    ):
        val = row[0].value
        if val:
            if "Count: " in val:
                employeeList[i].shift_count = getShiftCount(val)
                i += 1

    # initialize Employees' Hours
    employeeCount = 0
    for row in ws_schedule.iter_rows(
        min_col=column_index_from_string("H"),
        max_col=column_index_from_string("H"),
        min_row=3,
    ):
        val = row[0].value
        if j == employeeList[employeeCount].shift_count:
            employeeList[employeeCount].hours = float(val)
            j = 0
            employeeCount += 1
        elif val:
            j += 1
        else:
            continue

    # Initialize employees schedule
    for employee in employeeList:
        for colB, colC, colD, colE, colF, colG, colH in zip(
            ws_schedule.iter_cols(min_col=2, max_col=2, values_only=True),
            ws_schedule.iter_cols(min_col=3, max_col=3, values_only=True),
            ws_schedule.iter_cols(min_col=4, max_col=4, values_only=True),
            ws_schedule.iter_cols(min_col=5, max_col=5, values_only=True),
            ws_schedule.iter_cols(min_col=6, max_col=6, values_only=True),
            ws_schedule.iter_cols(min_col=7, max_col=7, values_only=True),
            ws_schedule.iter_cols(min_col=8, max_col=8, values_only=True),
        ):
            for B, C, D, E, F, G, H in zip(colB, colC, colD, colE, colF, colG, colH):
                if G:
                    if employee.last_name in G and employee.first_name in G:

                        employee.schedule.append(
                            Shift(
                                get_day_of_week(B),
                                convert_to_time(C),
                                convert_to_time(D),
                                float(H),
                                F,
                            )
                        )

    # getFacilityGuidelineRegular(ws_requirements)

    guidelineCheck(udm_regular_guideline, udm_supervisor_guideline, employeeList)
    filterList(employeeList)

    # Add Rows to new worksheet
    for employee in employeeList:
        ws_callsheet[f"A{k}"] = employee.last_name
        ws_callsheet[f"B{k}"] = employee.first_name
        ws_callsheet[f"C{k}"] = employee.shift_count
        ws_callsheet[f"D{k}"] = employee.hours
        ws_callsheet[f"E{k}"] = employee.supervisor
        ws_callsheet[f"F{k}"] = ", ".join(employee.missing_reqs)
        k += 1

        wb_requirementsCheck.save("SampleRequirementsCheckForBen.xlsx")


# Helper function to convert the day of the year into its respective day of the week
def get_day_of_week(date_obj):
    # date_obj = datetime.datetime.strptime(date_str, "%m/%d/%Y")
    day_of_week_num = date_obj.weekday()
    days_of_week = [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ]
    day_of_week = days_of_week[day_of_week_num]
    return day_of_week


# Helper function to return integer value of the column (Used to iterate through a column)
def column_index_from_string(column):
    index = 0
    for char in column.upper():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


# Retrieve last name from the full name of employee
def getLastName(employeeName):
    try:
        employeeName = employeeName[10:]
        last_name, first_name = employeeName.split(", ")
        return last_name
    except ValueError:
        return "Invalid format. Please use 'Last, First'."


# Retrieve first name from the full name of employee
def getFirstName(employeeName):
    try:
        employeeName = employeeName[10:]
        last_name, first_name = employeeName.split(", ")
        return first_name
    except ValueError:
        return "Invalid format. Please use 'Last, First'."


# Retrieve the number of shifts the employee is working
def getShiftCount(cellValue):
    stringNumber = cellValue[7:]
    return int(stringNumber)


def getFacilityGuidelineRegular(ws):
    req_list = []
    
    for row in ws.iter_rows(min_row=2):
        # Initialize column values from sheet
        row_number = row[0].row
        print(row_number)
        req_name = row[column_index_from_string("A") - 1].value
        req_type = row[column_index_from_string("B") - 1].value
        days = decodeDaysCheckboxCell(row_number, ws)
        min_start = row[column_index_from_string("D") - 1].value
        max_start = row[column_index_from_string("E") - 1].value
        min_end = row[column_index_from_string("F") - 1].value
        max_end = row[column_index_from_string("G") - 1].value
        shift_type = row[column_index_from_string("H") - 1].value
        cell_value = row[column_index_from_string("I") - 1].value
        exceptions = False
        if cell_value.lower() == "y":
            exceptions = True
        # TODO: Checkboxes for exceptions
        other_exception_shifts = row[column_index_from_string("K") - 1].value
        min_exception_shifts = row[column_index_from_string("L") - 1].value

        if req_type == "Day":
            day_req_type = DayRequirement(days)
            day_req = Requirement(day_req_type, exceptions, 1)
            req_list.append(day_req) 

        elif req_type == "Time":
            time_req_type = TimeRequirement(min_start, max_start, min_end, max_end)
            time_req = Requirement(time_req_type, exceptions, 1)
            req_list.append(time_req)

        elif req_type == "Shift Type":
            print()
            #TODO: Checkboxes for Shift Type

        else:
            print("Error: Invalid Input")

    return 0


def getFacilityGuidelineSupervisor():
    # TODO
    return 0


def decodeDaysCheckboxCell(row_number, ws):
    day_list = []
    row_number *= 10
    monday = ws.cell(row=row_number, column=3).value
    tuesday = ws.cell(row=row_number+1, column=3).value
    wednesday = ws.cell(row=row_number+2, column=3).value
    thursday = ws.cell(row=row_number+3, column=3).value
    friday = ws.cell(row=row_number+4, column=3).value
    saturday = ws.cell(row=row_number+5, column=3).value
    sunday = ws.cell(row=row_number+6, column=3).value
    na = ws.cell(row=row_number+7, column=3).value

    if na : return []
    
    if monday: day_list.append("Monday")
    if tuesday: day_list.append("Tuesday")
    if wednesday: day_list.append("Wednesday")
    if thursday: day_list.append("Thursday")
    if friday: day_list.append("Friday")
    if saturday: day_list.append("Saturday")
    if sunday: day_list.append("Sunday")
    
    
    return day_list
    

def guidelineCheck(guideline, sup_guideline, employeeList):
    for employee in employeeList:
        if employee.supervisor == "Yes":
            req_list = sup_guideline.requirement_list
            if sup_guideline.exceptions:
                exceptions = sup_guideline.exceptions
                req_excused = exceptions.requirement_excused
            else:
                exceptions = None
            min_hours = sup_guideline.min_hours

        else:
            req_list = guideline.requirement_list
            exceptions = guideline.exceptions
            req_excused = exceptions.requirement_excused
            min_hours = guideline.min_hours

        for req in req_list:
            req_type = req.requirement_type
            if not (meetsRequirement(employee, req_type, req.min_shifts, min_hours)):
                employee.schedule_valid = False

                # Check for excuses in the requirement
                if exceptions and req == req_excused:
                    if excuseRequirement(employee, exceptions, req):
                        employee.schedule_valid = True

    return employeeList


def excuseRequirement(employee, exceptions, req):
    req_list = (
        exceptions.excuse_req_list
    )  # If any of these requirements are met, it counts towards 1 shift in the min_shifts
    meets_req_count = 0
    min_shifts = exceptions.min_shifts

    for req in req_list:
        meets_req_count += countMetRequirements(employee, req.requirement_type)

    if meets_req_count >= min_shifts:
        return True

    else:
        return False


def countMetRequirements(employee, requirement_type):
    shift_count = 0

    if isinstance(requirement_type, DayRequirement):
        days = requirement_type.days_list
        for shift in employee.schedule:
            for day in days:
                if shift.day == day:
                    shift_count += 1

    elif isinstance(requirement_type, ShiftTypeRequirement):
        types = requirement_type.types_list
        for shift in employee.schedule:
            for shiftType in types:
                if shift.station == shiftType:
                    shift_count += 1

    elif isinstance(requirement_type, TimeRequirement):
        min_start = requirement_type.min_start
        min_end = requirement_type.min_end
        max_start = requirement_type.max_start
        max_end = requirement_type.max_end

        # Set necessary upper/lower bounds of times for values that are null
        if min_start == None:
            min_start = convert_to_time("12:00:00 AM")

        if max_start == None:
            max_start = convert_to_time("11:59:00 PM")

        if min_end == None:
            min_end = convert_to_time("12:00:00 AM")

        if max_end == None:
            max_end = convert_to_time("11:59:00 PM")

        # Sift through schedule and make sure at least 1 shift is in range
        for shift in employee.schedule:
            if (
                shift.start_time >= min_start
                and shift.start_time <= max_start
                and shift.end_time >= min_end
                and shift.end_time <= max_end
            ):
                shift_count += 1

    return shift_count


def meetsRequirement(employee, requirement_type, min_shifts, min_hours):
    shift_count = 0
    meetsRequirement = False

    if isinstance(requirement_type, DayRequirement):
        days = requirement_type.days_list
        for shift in employee.schedule:
            for day in days:
                if shift.day == day:
                    shift_count += 1
                    if shift_count == min_shifts:
                        meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append("No Weekend")

    elif isinstance(requirement_type, ShiftTypeRequirement):
        types = requirement_type.types_list
        if "Supervisor" in types:
            min_shifts = employee.shift_count

        for shift in employee.schedule:
            for shiftType in types:
                if shift.station == shiftType:
                    shift_count += 1
                    # print("Shift Type Requirement Fulfilled - " + employee.first_name + " " + employee.last_name)
                    if shift_count == min_shifts:
                        meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append("Shift Type")

    elif isinstance(requirement_type, TimeRequirement):
        min_start = requirement_type.min_start
        min_end = requirement_type.min_end
        max_start = requirement_type.max_start
        max_end = requirement_type.max_end

        # Set necessary upper/lower bounds of times for values that are null
        if min_start == None:
            min_start = convert_to_time("12:00:00 AM")

        if max_start == None:
            max_start = convert_to_time("11:59:00 PM")

        if min_end == None:
            min_end = convert_to_time("12:00:00 AM")

        if max_end == None:
            max_end = convert_to_time("11:59:00 PM")

        # Sift through schedule and make sure at least 1 shift is in range
        for shift in employee.schedule:
            if (
                shift.start_time >= min_start
                and shift.start_time <= max_start
                and shift.end_time >= min_end
                and shift.end_time <= max_end
            ):
                shift_count += 1
                # print("Time Requirement Fulfilled - " + employee.first_name + " " + employee.last_name)
                if shift_count == min_shifts:
                    meetsRequirement = True
        if not (meetsRequirement):
            employee.missing_reqs.append("Time Req not met")

    if employee.hours < min_hours:
        meetsRequirement = False
        if not ("Under Hours" in employee.missing_reqs):
            employee.missing_reqs.append("Under Hours")

    return meetsRequirement


def filterList(employeeList):
    for employee in employeeList[:]:  # Iterate over a shallow copy of the list
        if employee.schedule_valid:
            employeeList.remove(employee)


def convert_to_time(time_input):
    # Check if the input is a datetime object
    if isinstance(time_input, datetime):
        datetime_obj = time_input
    else:
        # Define the datetime formats
        full_datetime_format = "%m/%d/%Y %I:%M:%S %p"
        time_only_format = "%I:%M:%S %p"

        try:
            # Try to parse the full date-time string
            datetime_obj = datetime.strptime(time_input, full_datetime_format)
        except ValueError:
            # If that fails, try to parse the time-only string
            datetime_obj = datetime.strptime(time_input, time_only_format)

    # Extract the time component from the datetime object
    time_struct = datetime_obj.time()

    return time_struct


if __name__ == "__main__":
    main()
